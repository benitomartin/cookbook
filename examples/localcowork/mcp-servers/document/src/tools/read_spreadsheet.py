"""
document.read_spreadsheet — Read data from an Excel or CSV spreadsheet.

Non-destructive: executes immediately, no confirmation needed.
Supports .xlsx, .xls, .csv, .tsv files.
"""

from __future__ import annotations

import csv
import os
from pathlib import Path
from typing import Any

from pydantic import BaseModel, Field

from mcp_base import MCPError, MCPResult, MCPTool, ErrorCodes
from validation import assert_sandboxed, assert_absolute_path


class Params(BaseModel):
    """Parameters for read_spreadsheet."""

    path: str = Field(description="Path to spreadsheet")
    sheet: str | None = Field(default=None, description="Sheet name (for multi-sheet files)")


class Result(BaseModel):
    """Return value for read_spreadsheet."""

    headers: list[str]
    rows: list[dict[str, Any]]
    total_rows: int


class ReadSpreadsheet(MCPTool[Params, Result]):
    """Read data from an Excel or CSV spreadsheet."""

    name = "document.read_spreadsheet"
    description = "Read data from an Excel or CSV spreadsheet"
    confirmation_required = False
    undo_supported = False

    async def execute(self, params: Params) -> MCPResult[Result]:
        """Read and return spreadsheet data."""
        assert_absolute_path(params.path, "path")
        assert_sandboxed(params.path)

        if not os.path.exists(params.path):
            raise MCPError(ErrorCodes.FILE_NOT_FOUND, f"File not found: {params.path}")

        ext = Path(params.path).suffix.lower()

        try:
            if ext in (".xlsx", ".xls"):
                return MCPResult(success=True, data=_read_excel(params.path, params.sheet))
            if ext == ".csv":
                return MCPResult(success=True, data=_read_csv(params.path, ","))
            if ext == ".tsv":
                return MCPResult(success=True, data=_read_csv(params.path, "\t"))

            raise MCPError(
                ErrorCodes.INVALID_PARAMS,
                f"Unsupported spreadsheet format: {ext}. Supported: .xlsx, .xls, .csv, .tsv",
            )

        except MCPError:
            raise
        except Exception as e:
            raise MCPError(ErrorCodes.INTERNAL_ERROR, f"Failed to read spreadsheet: {e}") from e


def _read_excel(file_path: str, sheet_name: str | None) -> Result:
    """Read an Excel file using openpyxl."""
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)

    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    if ws is None:
        raise MCPError(ErrorCodes.INTERNAL_ERROR, "No active worksheet found")

    all_rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    wb.close()

    if not all_rows:
        return Result(headers=[], rows=[], total_rows=0)

    # First row is headers
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(all_rows[0])]
    rows = [dict(zip(headers, row)) for row in all_rows[1:]]

    return Result(headers=headers, rows=rows, total_rows=len(rows))


def _read_csv(file_path: str, delimiter: str) -> Result:
    """Read a CSV/TSV file."""
    with open(file_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        headers = reader.fieldnames or []
        rows = list(reader)

    return Result(headers=list(headers), rows=rows, total_rows=len(rows))
